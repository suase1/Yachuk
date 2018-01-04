from django.shortcuts import render
from urllib.request import urlopen
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook
import os
from django.utils import timezone

# Create your views here.
def login(request):
    # posts = Post.objects.filter(published_date__lte=timezone.now()).order_by('published_date')
    return render(request, 'inventory/login.html', {})

def main(request):
    return render(request, 'inventory/main.html', {})
#<-------------------------------------------------------------------------------------------------->


# <-----------------------------  yachuk  ------------------------------>
def yachuk_get_CodeNumber():
    CodeNumber=[]
    compare=re.compile("\([A-Z]*[0-9]+\)")
    compare_only_code = re.compile("[A-Z]*[0-9]+")
    for i in range(1,10):
        try:
            startpage = "http://ninetofive.cafe24.com/product/list.html?cate_no=130&page="
            page = startpage + str(i)
            html = urlopen(page)
            bs0bj = BeautifulSoup(html, "html.parser")
            for link in bs0bj.find("div",{"class":"xans-element- xans-product xans-product-listnormal"}).find_all("a",{"class":"name"}):
                try:
                    codenumber1 = compare.search(str(link)).group()
                    codenumber = compare_only_code.search(codenumber1).group().strip()
                    CodeNumber.append(codenumber)
                except AttributeError: pass
        except : break
    return CodeNumber

# <-----------------------------  ssaka  ------------------------------>
def ssaka_get_CodeNumber(yachuk_codelist):
    Accordance_CodeNumber = []
    Accordance_PR_link = []
    compare = re.compile("\([0-9]+\)")
    compare_only_code = re.compile("[0-9]+")
    head_page = "http://www.ssaka.co.kr/product/pro_list?page="
    tail_page = "&cate_re=10001&cate0=&cate1=1&cate3=13&order=item_serial&cate_re=10001&order=item_serial&limit=36&go_list=Y&probrand=&prosailo=&prosize=&prouse=&str_color="
    for i in range(1,15):
        try:
            page = head_page + str(i) + tail_page
            html = urlopen(page)
            bs0bj = BeautifulSoup(html, "html.parser")

            for code in bs0bj.find("div", {"id":"sProlistArea"}).find_all("a"):
                try:
                    n1 = compare.search(code.contents[1].attrs['title'])
                    n2 = compare_only_code.search(n1.group())
                    accordance_codenumber = n2.group().strip()
                    if accordance_codenumber in yachuk_codelist:
                        accordance_link = "http://www.ssaka.co.kr"+code.attrs['href']
                        Accordance_CodeNumber.append(accordance_codenumber)
                        Accordance_PR_link.append(accordance_link)
                except : pass
        except : break
    return Accordance_PR_link,Accordance_CodeNumber



def ssaka_Inventory(accordance_link, accordance_codenumber):
    accordance = list(zip(accordance_link, accordance_codenumber))
    inventory = {}
    for link, code in accordance:
        html = urlopen(link)
        bs0bj = BeautifulSoup(html, "html.parser")
        inventory[code] = {}
        size_info = [size for size in bs0bj.find_all('input', {"name":"SIZE[]"})]
        count_info = [count for count in bs0bj.find_all('input', {"name":"ex_qty[]"})]
        sizes = [size.attrs['value'].strip() for size in size_info]
        counts = [count.attrs['value'].strip() for count in count_info]
        inventory_info = list(zip(sizes, counts))
        for size, count in inventory_info:
            inventory[code][size] =count
    return inventory

# <-----------------------------  fifa  ------------------------------>
def fifa_get_CodeNumber(yachuk_codelist):
    Accordance_CodeNumber = []
    Accordance_PR_link = []
    compare = re.compile("[A-Z]*[0-9]+")

    for i in range(1,18):
        startpage = "http://fifas.co.kr/shop/goods/goods_list.php?category=039&brand=1&page="
        page = startpage + str(i)
        html = urlopen(page)
        bs0bj = BeautifulSoup(html, "html.parser")
        for link in bs0bj.find_all("a", {"class":"pname"}):
            n = compare.search(link.string)
            m = n.group().strip()
            if m in yachuk_codelist:
                Accordance_CodeNumber.append(m)
                pr_link = link.attrs['href'].split('..')[1]
                Accordance_PR_link.append("http://fifas.co.kr/shop" + pr_link)
            else: pass
    return Accordance_PR_link, Accordance_CodeNumber


def fifa_Inventory(accordance_link, accordance_codenumber):
    accordance = list(zip(accordance_link, accordance_codenumber))
    inventory = {}
    for link, code in accordance:
        html = urlopen(link)
        bs0bj = BeautifulSoup(html, "html.parser")
        inventory[code] = {}
        goodsno = bs0bj.find('input', {"name": "goodsno"}).attrs['value']
        size_info = [size for size in bs0bj.find_all('input', {"name":"multi_opt["+str(goodsno)+"][]"})]
        count_info = [count for count in bs0bj.find_all('input', {"name":"multi_stock["+str(goodsno)+"][]"})]
        sizes = [size.attrs['value'].strip() for size in size_info]
        counts = [count.attrs['value'].strip() for count in count_info]
        inventory_info = list(zip(sizes, counts))
        for size, count in inventory_info:
            inventory[code][size] =count

    return inventory



# <-----------------------------  kika  ------------------------------>
def kika_get_CodeNumber(yachuk_codelist):
    Accordance_CodeNumber = []
    Accordance_PR_link = []
    compare = re.compile("\([A-Z]*[0-9]+\)")
    compare_only_code = re.compile("[A-Z]*[0-9]+")

    for i in range(1,15):
        startpage = "http://www.aaasports.co.kr/front/productlist.php?code=001000000000&brandcode=2&listnum=30&sort=&block=0&gotopage="
        page = startpage + str(i)
        html = urlopen(page)
        bs0bj = BeautifulSoup(html, "html.parser")

        for tag in bs0bj.find_all(attrs={"class":"mainprname"}):
            n1=compare.search(tag.string)
            n2=compare_only_code.search(n1.group()).group().strip()
            if n2 in yachuk_codelist:
                Accordance_CodeNumber.append(n2)

                pr_href = tag.parent.attrs['href']
                pr_link = pr_href.split('..')[1]
                result = "http://www.aaasports.co.kr" + pr_link
                Accordance_PR_link.append(result)
    return Accordance_PR_link, Accordance_CodeNumber


def kika_Inventory(accordance_link, accordance_codenumber):
    accordance = list(zip(accordance_link, accordance_codenumber))
    inventory = {}
    for link, code in accordance:
        html = urlopen(link)
        bs0bj = BeautifulSoup(html, "html.parser")
        inventory[code] = {}

        size_info = [size for size in bs0bj.find_all('input', {"name":"ord_spec"})]
        count_info = [count for count in bs0bj.find_all('input', {"name":"kika_quantity"})]
        sizes = [size.attrs['value'][-4:].strip() for size in size_info]    #use of slicing for erasing color category term
        counts = [count.attrs['value'].strip() for count in count_info]

        inventory_info = list(zip(sizes, counts))
        for size, count in inventory_info:
            inventory[code][size] =count

    return inventory


# <--------------------integration------------------->
def comparison(ssaka_inventory, fifa_inventory, kika_inventory):
    Wholesale = {**ssaka_inventory, **fifa_inventory, **kika_inventory}
    print(Wholesale)
    desktoppath = os.path.expanduser('~')
    yachuk_excel = desktoppath + "\\Desktop\\inventory_information\\ninetofive.xlsx"
    wb_Yachuk = load_workbook(yachuk_excel)
    ws_Yachuk = wb_Yachuk.active

    for i in range(2,len(list(ws_Yachuk.rows))+1):
        try:
            ws_Yachuk.cell(row=i, column=10).value = Wholesale[str(ws_Yachuk.cell(row=i, column=2).value)][str(ws_Yachuk.cell(row=i, column=7).value)]
        except KeyError:
            ws_Yachuk.cell(row=i, column=10).value = 0

    desktoppath = os.path.expanduser('~')
    current_time = timezone.now()
    wb_Yachuk.save(desktoppath + "\\Desktop\\inventory_information\\Integration_version_" + str(current_time)[:11] + ".xlsx")
    return 0
#<--------------------------------------------------------------------------------------->
def crawling(request):
    # <---------------Implementation------------->
    yachuk_codelist = yachuk_get_CodeNumber()

    ssaka_link, ssaka_code = ssaka_get_CodeNumber(yachuk_codelist)
    ssaka_inventory = ssaka_Inventory(ssaka_link, ssaka_code)

    fifa_link, fifa_code = fifa_get_CodeNumber(yachuk_codelist)
    fifa_inventory = fifa_Inventory(fifa_link,fifa_code)

    kika_link, kika_code = kika_get_CodeNumber(yachuk_codelist)
    kika_inventory = kika_Inventory(kika_link,kika_code)


    # <--------------------  integration  -------------------->
    comparison(ssaka_inventory,fifa_inventory,kika_inventory)
    return render(request, 'inventory/work_end.html', {})
